import yt_dlp
import os
import requests
import time
from openpyxl import Workbook, load_workbook
import math, openpyxl
import re
import datetime
import random
import string
import shutil
import pyzipper  # 必须安装: pip install pyzipper

# ================= 配置区域 =================
# 1. FFmpeg 路径
FFMPEG_PATH = r'C:\ffmpeg\bin' 

# 2. 初始目标链接 (会被下方的主程序循环覆盖)
TARGET_URL = "https://www.bilibili.com/video/BV1DLznBgERM/"

# 3. 根目录
BASE_DIR = "File2"

# 4. 伪装头
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://www.bilibili.com"
}
# ===========================================

def _ReadXlsl(file_path):
    print(f"正在快速参考文件: {file_path} ")
    if not os.path.exists(file_path):
        return []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        all_sheets = wb.sheetnames
        if not all_sheets: return []
        ws = wb[all_sheets[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

def check_env():
    if not os.path.exists(os.path.join(FFMPEG_PATH, 'ffmpeg.exe')):
        print(f"❌ 错误：在 {FFMPEG_PATH} 找不到 ffmpeg.exe")
        return False
    if not os.path.exists(BASE_DIR):
        os.makedirs(BASE_DIR)
    return True

def get_next_index_from_excel(base_dir):
    excel_path = os.path.join(base_dir, "download_report.xlsx")
    max_idx = 0
    if not os.path.exists(excel_path):
        if not os.path.exists(base_dir):
            return 1
        for folder_name in os.listdir(base_dir):
            if folder_name.isdigit():
                idx = int(folder_name)
                if idx > max_idx:
                    max_idx = idx
        return max_idx + 1

    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None and isinstance(row[0], int):
                if row[0] > max_idx:
                    max_idx = row[0]
        wb.close()
    except Exception as e:
        print(f"读取Excel索引出错，降级为扫描文件夹: {e}")
        return 1
    return max_idx + 1

def generate_password(length=6):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

def create_encrypted_zip(source_dir, output_zip_path, password):
    with pyzipper.AESZipFile(output_zip_path, 'w', compression=pyzipper.ZIP_LZMA, encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(password.encode('utf-8'))
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                if file == os.path.basename(output_zip_path):
                    continue
                file_path = os.path.join(root, file)
                arcname = file 
                zf.write(file_path, arcname)

def format_file_size(size_bytes):
    if not size_bytes: return "0B"
    size_name = ("B", "KB", "MB", "GB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    return f"{round(size_bytes / p, 2)} {size_name[i]}"

def format_seconds(seconds):
    if not seconds: return "00:00:00"
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "{:02d}:{:02d}:{:02d}".format(int(h), int(m), int(s))

def format_date_str(date_str):
    if date_str and len(date_str) == 8:
        return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    return date_str

def get_bilibili_extra_info(bvid):
    info_data = {"view": 0, "favorite": 0}
    related_list = []
    try:
        api_view = f"https://api.bilibili.com/x/web-interface/view?bvid={bvid}"
        resp = requests.get(api_view, headers=HEADERS, timeout=5)
        data = resp.json()
        if data['code'] == 0:
            stat = data['data']['stat']
            info_data['view'] = stat.get('view', 0)
            info_data['favorite'] = stat.get('favorite', 0)
    except Exception:
        pass

    try:
        api_related = f"https://api.bilibili.com/x/web-interface/archive/related?bvid={bvid}"
        resp = requests.get(api_related, headers=HEADERS, timeout=5)
        data = resp.json()
        if data['code'] == 0:
            for item in data['data']:
                r_bvid = item.get('bvid')
                if r_bvid:
                    clean_url = f"https://www.bilibili.com/video/{r_bvid}"
                    related_list.append({
                        "title": item.get('title', '未知'),
                        "owner": item.get('owner', {}).get('name', '未知'),
                        "view": item.get('stat', {}).get('view', 0),
                        "url": clean_url
                    })
    except Exception:
        pass
    return info_data, related_list

def save_related_excel(save_path, related_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "相关推荐"
    ws.append(["视频标题", "UP主", "播放量", "链接"])
    for item in related_data:
        ws.append([item['title'], item['owner'], item['view'], item['url']])
    wb.save(save_path)

# === 专门用于处理重试任务的函数（已修改：覆盖旧数据） ===
def process_retry_task(retry_index, retry_url):
    """
    针对指定的 INDEX 和 URL 进行重新下载
    逻辑：清空目录 -> 下载 -> 压缩 -> 【覆盖Excel旧行】
    """
    if not check_env(): return

    real_index = str(retry_index)
    video_dir = os.path.join(BASE_DIR, real_index)

    print(f"\n♻️ [重试模式] 正在修正 Index: {real_index} | URL: {retry_url}")

    # 1. 文件夹处理逻辑：强制清空重建
    if os.path.exists(video_dir):
        print(f"   🧹 目录已存在，正在清空: {video_dir}")
        try:
            shutil.rmtree(video_dir)
            time.sleep(1)
            os.makedirs(video_dir)
        except Exception as e:
            print(f"   ❌ 无法删除旧目录，请检查文件占用: {e}")
            return
    else:
        print(f"   📂 新建目录: {video_dir}")
        os.makedirs(video_dir)

    # 2. 准备变量
    excel_path = os.path.join(BASE_DIR, "download_report.xlsx")
    today_date = datetime.datetime.now().strftime("%Y-%m-%d")
    
    # 默认值
    video_title = 'Unknown Title'
    uploader = "未知作者"
    upload_date_str = ""
    duration_str = "00:00:00"
    size_str = "0B"
    zip_password = ""
    stats_view = 0
    stats_fav = 0

    # 3. API 获取信息
    bv_match = re.search(r'(BV\w+)', retry_url)
    bvid = bv_match.group(1) if bv_match else None
    if bvid:
        stats_info, related_videos = get_bilibili_extra_info(bvid)
        stats_view = stats_info['view']
        stats_fav = stats_info['favorite']
        if related_videos:
            save_related_excel(os.path.join(video_dir, "related_videos.xlsx"), related_videos)

    # 4. 下载配置
    final_file_path = os.path.join(video_dir, f"{real_index}.mp4")
    ydl_opts = {
        'ffmpeg_location': FFMPEG_PATH,
        'outtmpl': f'{video_dir}/{real_index}.%(ext)s',
        'format': 'bv[vcodec^=avc][height<=1080]+ba/b[height<=1080]',
        'merge_output_format': 'mp4',
        'writethumbnail': True,
        'postprocessors': [
            {'key': 'FFmpegThumbnailsConvertor', 'format': 'png'},
            {'key': 'FFmpegVideoConvertor', 'preferedformat': 'mp4'}
        ],
        'quiet': False,
        'no_warnings': True,
        'ignoreerrors': True,
    }

    # 5. 执行下载
    download_success = False
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl_worker:
            meta = ydl_worker.extract_info(retry_url, download=True)
            if meta:
                video_title = meta.get('title', video_title)
                duration_str = format_seconds(meta.get('duration', 0))
                uploader = meta.get('uploader', '未知作者')
                upload_date_str = format_date_str(meta.get('upload_date'))
                description = meta.get('description', '无简介')
                
                desc_path = os.path.join(video_dir, "简介.txt")
                with open(desc_path, "w", encoding="utf-8") as f:
                    f.write(description if description else "无简介")
                download_success = True
    except Exception as e:
        print(f"   ❌ 重试下载失败: {e}")

    # 6. 压缩打包
    if download_success:
        if os.path.exists(final_file_path):
            size_str = format_file_size(os.path.getsize(final_file_path))
        
        time.sleep(2) 
        
        try:
            zip_password = generate_password(6)
            zip_filename = f"{real_index}.zip"
            zip_full_path = os.path.join(video_dir, zip_filename)
            
            print(f"   🔒 正在加密压缩 (密码: {zip_password})...")
            create_encrypted_zip(video_dir, zip_full_path, zip_password)
            
            for f_name in os.listdir(video_dir):
                if f_name != zip_filename:
                    try:
                        f_path = os.path.join(video_dir, f_name)
                        if os.path.isdir(f_path): shutil.rmtree(f_path)
                        else: os.remove(f_path)
                    except: pass
        except Exception as e:
            print(f"   ⚠️ 打包失败: {e}")
            zip_password = "打包出错"

    # 7. 更新 Excel (覆盖旧行模式)
    try:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Index", "视频名称", "作者", "发布日期", "下载日期", "时长", "大小", "播放量", "收藏数", "解压密码", "原始链接"])
        
        # 准备好要写入的数据（无论成功失败都覆盖）
        if download_success:
            row_data = [
                int(real_index), video_title, uploader, upload_date_str, today_date,
                duration_str, size_str, stats_view, stats_fav, zip_password, retry_url
            ]
        else:
            row_data = [
                int(real_index), video_title, "Error", "", today_date, 
                "", "0B", 0, 0, "", retry_url
            ]

        # === 核心逻辑修改：查找并覆盖 ===
        target_row_obj = None
        # 遍历第一列查找 Index
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value == int(real_index):
                target_row_obj = row[0]
                break
        
        if target_row_obj:
            # 如果找到了，就原地修改这一行
            row_num = target_row_obj.row
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=val)
            print(f"   --> Excel旧数据已覆盖 (Index: {real_index})")
        else:
            # 如果没找到（极少情况），则追加
            ws.append(row_data)
            print(f"   --> Index未找到，已追加新数据 (Index: {real_index})")
            
        wb.save(excel_path)
    except Exception as e:
        print(f"   ❌ Excel写入失败: {e}")


# 原有的批量下载逻辑（常规逻辑，只追加）
def process_download():
    if not check_env(): return
    excel_path = os.path.join(BASE_DIR, "download_report.xlsx")
    columns = ["Index", "视频名称", "作者", "发布日期", "下载日期", "时长", "大小", "播放量", "收藏数", "解压密码", "原始链接"]
    
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "资源列表"
        ws.append(columns)

    current_index = get_next_index_from_excel(BASE_DIR)
    
    print("🔍 正在解析初始信息...")
    with yt_dlp.YoutubeDL({'quiet': True, 'extract_flat': True}) as ydl:
        try:
            info = ydl.extract_info(TARGET_URL, download=False)
        except Exception as e:
            print(f"❌ 解析失败: {e}")
            return

    video_list = info['entries'] if 'entries' in info else [info]
    today_date = datetime.datetime.now().strftime("%Y-%m-%d")

    for i, item in enumerate(video_list):
        real_index = str(current_index + i)
        origin_url = item.get('url', item.get('webpage_url', TARGET_URL))
        video_title = item.get('title', 'Unknown Title')
        uploader = "未知作者"
        upload_date_str = ""
        duration_str = "00:00:00"
        size_str = "0B"
        zip_password = ""
        stats_view = 0
        stats_fav = 0
        
        print(f"\n[{i+1}/{len(video_list)}] 处理: {video_title} (Index: {real_index})")

        video_dir = os.path.join(BASE_DIR, real_index)
        if not os.path.exists(video_dir):
            os.makedirs(video_dir)

        bv_match = re.search(r'(BV\w+)', origin_url)
        bvid = bv_match.group(1) if bv_match else None
        if bvid:
            stats_info, related_videos = get_bilibili_extra_info(bvid)
            stats_view = stats_info['view']
            stats_fav = stats_info['favorite']
            if related_videos:
                save_related_excel(os.path.join(video_dir, "related_videos.xlsx"), related_videos)

        final_file_path = os.path.join(video_dir, f"{real_index}.mp4")
        ydl_opts = {
            'ffmpeg_location': FFMPEG_PATH,
            'outtmpl': f'{video_dir}/{real_index}.%(ext)s',
            'format': 'bv[vcodec^=avc][height<=1080]+ba/b[height<=1080]',
            'merge_output_format': 'mp4',
            'writethumbnail': True,
            'postprocessors': [
                {'key': 'FFmpegThumbnailsConvertor', 'format': 'png'},
                {'key': 'FFmpegVideoConvertor', 'preferedformat': 'mp4'}
            ],
            'quiet': False,
            'no_warnings': True,
            'ignoreerrors': True,
        }

        download_success = False
        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl_worker:
                meta = ydl_worker.extract_info(origin_url, download=True)
                if meta:
                    video_title = meta.get('title', video_title)
                    duration_str = format_seconds(meta.get('duration', 0))
                    uploader = meta.get('uploader', '未知作者')
                    upload_date_str = format_date_str(meta.get('upload_date'))
                    description = meta.get('description', '无简介')
                    
                    desc_path = os.path.join(video_dir, "简介.txt")
                    with open(desc_path, "w", encoding="utf-8") as f:
                        f.write(description if description else "无简介")
                    download_success = True
        except Exception as e:
            print(f"   ❌ 下载阶段出错: {e}")

        if download_success:
            if os.path.exists(final_file_path):
                size_str = format_file_size(os.path.getsize(final_file_path))
            time.sleep(2) 
            try:
                zip_password = generate_password(6)
                zip_filename = f"{real_index}.zip"
                zip_full_path = os.path.join(video_dir, zip_filename)
                print(f"   🔒 正在加密压缩...")
                create_encrypted_zip(video_dir, zip_full_path, zip_password)
                for f_name in os.listdir(video_dir):
                    if f_name != zip_filename:
                        try:
                            f_path = os.path.join(video_dir, f_name)
                            if os.path.isdir(f_path): shutil.rmtree(f_path)
                            else: os.remove(f_path)
                        except: pass
            except Exception as e:
                print(f"   ⚠️ 打包失败: {e}")
                zip_password = "打包出错-文件未加密"

        if download_success:
            row_data = [
                int(real_index), video_title, uploader, upload_date_str, today_date,
                duration_str, size_str, stats_view, stats_fav, zip_password, origin_url
            ]
            ws.append(row_data)
            wb.save(excel_path)
            print(f"   --> Excel 更新完毕 (Index: {real_index})")
        else:
            ws.append([int(real_index), video_title, "Error", "", today_date, "", "0B", 0, 0, "", origin_url])
            wb.save(excel_path)
            
        time.sleep(2)

    print(f"\n🎉 任务完成！")


if __name__ == '__main__':
    # 1. 优先处理之前下载失败（0B）的任务
    print("\n🔍 正在检查是否有需要修复的 0B 视频...")
    
    # 扫描列表（先全部读取到内存，防止边读边写出问题）
    retry_list = []
    
    if os.path.exists("File2/download_report.xlsx"):
        _dat = _ReadXlsl("File2/download_report.xlsx")
        if _dat and len(_dat) > 1:
            _dat = _dat[1:] # 跳过表头
            for obj in _dat:
                # 检查 Index(0) 和 URL(10) 和 大小(6)
                if len(obj) > 10 and obj[6] == '0B':
                    retry_list.append((obj[0], obj[10]))

    if retry_list:
        print(f"⚠️ 发现 {len(retry_list)} 个 0B 视频，开始尝试修复（覆盖模式）...")
        print("="*40)
        input("Start")
        for idx, retry_url in retry_list:
            if not retry_url: continue # 防止空链接
            process_retry_task(idx, retry_url)
            time.sleep(3) 
            
        print("="*40)
        print("✅ 0B 修复流程结束。\n")
    else:
        print("✅ 未发现 0B 视频 或 文件不存在，无需修复。\n")

    # 2. 常规任务
    print("🚀 准备开始常规 BID.xlsx 任务...")
    input("Start")
    if os.path.exists("BID.xlsx"):
        _data = _ReadXlsl("BID.xlsx")[1:]
        _data.reverse()
        _data = _data[1500:] 
        
        print(f"常规任务长度: {len(_data)}")
        if _data:
            input("按回车开始...")
            i = 1
            for _burl in _data:
                if not _burl or len(_burl) < 2 or not _burl[1]: continue
                TARGET_URL = _burl[1] 
                print(f"\n========================================")
                print(f"常规进度 {i}/{len(_data)} | 目标: {TARGET_URL}")
                
                process_download()
                
                time.sleep(6)
                print("\a")
                i += 1
    else:
        print("❌ 未找到 BID.xlsx，无法执行常规任务。")